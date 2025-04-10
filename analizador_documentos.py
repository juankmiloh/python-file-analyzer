# -*- coding: utf-8 -*-
# Análisis de documentos para búsqueda de palabras clave (con persistencia)

import os
import logging
import argparse
import openpyxl
import unicodedata
import concurrent.futures
import re
import textract
import docx
import fitz  # PyMuPDF
from tqdm import tqdm

logging.getLogger("PyPDF2").setLevel(logging.ERROR)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("analizador.log"),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

class ExtractionError(Exception):
    def __init__(self, mensaje, archivo=None):
        super().__init__(mensaje)
        self.archivo = archivo

def normalizar(texto):
    return unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII").lower()

def extraer_texto_pdf(path):
    try:
        with fitz.open(path) as doc:
            return " ".join(page.get_text() for page in doc)
    except Exception as e:
        raise ExtractionError(f"Error leyendo PDF con PyMuPDF: {e}", archivo=path)

def extraer_texto_doc(path):
    try:
        return textract.process(path).decode('utf-8')
    except Exception as e:
        raise ExtractionError(f"Error leyendo DOC: {e}", archivo=path)

def extraer_texto_docx(path):
    try:
        doc = docx.Document(path)
        return " ".join(paragraph.text for paragraph in doc.paragraphs)
    except Exception as e:
        raise ExtractionError(f"Error leyendo DOCX: {e}", archivo=path)

def extraer_texto_xlsx(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        text = ""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        text += str(cell.value) + " "
        return text
    except Exception as e:
        raise ExtractionError(f"Error leyendo XLSX: {e}", archivo=path)

def procesar_archivo(path, incluir_resumen):
    nombre = os.path.basename(path)
    extension = os.path.splitext(nombre)[1].lower().lstrip('.')

    # 🚫 Validar tamaño del archivo (en bytes)
    max_size_mb = 23
    max_size_bytes = max_size_mb * 1024 * 1024
    if os.path.getsize(path) > max_size_bytes:
        logger.warning(f"Archivo ignorado por exceder los {max_size_mb}MB: {nombre}")
        return None

    texto = ""
    tipo = extension.upper()

    try:
        if extension == "pdf":
            texto = extraer_texto_pdf(path)
        elif extension == "docx":
            texto = extraer_texto_docx(path)
        elif extension == "doc":
            texto = extraer_texto_doc(path)
        elif extension == "xlsx":
            texto = extraer_texto_xlsx(path)
        elif extension in ["txt", "csv"]:
            with open(path, 'r', encoding="utf-8", errors="ignore") as f:
                texto = f.read()
        else:
            return None
    except ExtractionError as e:
        logger.error(f"{e} (Archivo: {e.archivo})")
        return None

    texto_normalizado = normalizar(texto)
    coincidencias = {
        palabra: len(re.findall(rf"\b{re.escape(normalizar(palabra))}\b", texto_normalizado))
        for palabra in PALABRAS_CLAVE
    }
    total = sum(coincidencias.values())

    resultado = {
        "ruta": path,
        "nombre": nombre,
        "tipo": tipo,
        "coincidencias": total,
        "detalle": coincidencias
    }

    if incluir_resumen:
        frases = re.split(r'(?<=[\.\n])\s+', texto)
        resumen_por_palabra = {}

        for palabra in PALABRAS_CLAVE:
            frases_encontradas = []
            palabra_normalizada = normalizar(palabra)

            for frase in frases:
                frase_normalizada = normalizar(frase)
                if re.search(rf"\b{re.escape(palabra_normalizada)}\b", frase_normalizada):
                    frases_encontradas.append(frase.strip())

            if frases_encontradas:
                resumen_por_palabra[palabra] = frases_encontradas

        resultado["resumen"] = resumen_por_palabra or {"sin_resultados": ["No se encontraron frases con palabras clave."]}

    return resultado

def cargar_archivos_procesados():
    procesados = set()
    if os.path.exists("procesados.txt"):
        with open("procesados.txt", "r", encoding="utf-8") as f:
            procesados = set(line.strip() for line in f)
    return procesados

def guardar_archivos_procesados(paths):
    with open("procesados.txt", "a", encoding="utf-8") as f:
        for path in paths:
            f.write(f"{path}\n")

def cargar_resultados_previos():
    resultados = []
    if not os.path.exists("informe_documentos.xlsx"):
        print("No se encontró el archivo informe_documentos.xlsx. Se generará uno nuevo.")
        return resultados
    
    print("Cargando resultados previos de informe_documentos.xlsx...")

    wb = openpyxl.load_workbook("informe_documentos.xlsx")
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        ruta = row[0]
        nombre = row[1]
        tipo = row[2]
        coincidencias = row[3]
        detalle = {PALABRAS_CLAVE[i]: row[4 + i] for i in range(len(PALABRAS_CLAVE))}
        resumen = {}
        if len(row) > 4 + len(PALABRAS_CLAVE) and row[4 + len(PALABRAS_CLAVE)]:
            resumen["resumen"] = [row[4 + len(PALABRAS_CLAVE)]]

        resultados.append({
            "ruta": ruta,
            "nombre": nombre,
            "tipo": tipo,
            "coincidencias": coincidencias,
            "detalle": detalle,
            "resumen": resumen
        })

    return resultados

def analizar_documentos(ruta, incluir_resumen, forzar_reprocesamiento=False):
    resultados_nuevos = []
    archivos = []
    procesados = set()
    if not forzar_reprocesamiento:
        procesados = cargar_archivos_procesados()

    for root, dirs, files in os.walk(ruta):
        for file in files:
            path = os.path.abspath(os.path.join(root, file))
            if path not in procesados:
                archivos.append(path)

    logger.info(f"{len(archivos)} archivos nuevos por procesar.")
    nuevos_procesados = []

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futuras = {executor.submit(procesar_archivo, path, incluir_resumen): path for path in archivos}
        for future in tqdm(concurrent.futures.as_completed(futuras), total=len(futuras), desc="Procesando documentos"):
            tqdm.write(f"🔄 Procesando archivo: {futuras[future]}")
            path = futuras[future]
            try:
                resultado = future.result()
                if resultado:
                    resultados_nuevos.append(resultado)
                    nuevos_procesados.append(path)
            except Exception as e:
                logger.error(f"Error procesando archivo {path}: {e}")

    guardar_archivos_procesados(nuevos_procesados)

    return resultados_nuevos

def guardar_txt(resultados, incluir_resumen):
    with open("informe_documentos.txt", "w", encoding="utf-8") as f:
        for r in resultados:
            ruta_completa = os.path.abspath(r["ruta"])
            ruta_directorio = os.path.dirname(ruta_completa)
            nombre_archivo = r["nombre"]
            f.write(f"Ruta: {ruta_directorio}\n")
            f.write(f"Archivo: {nombre_archivo} ({r['tipo']})\n")
            f.write(f"Total coincidencias: {r['coincidencias']}\n")
            for palabra, cantidad in r['detalle'].items():
                f.write(f"   {palabra}: {cantidad}\n")
            if incluir_resumen:
                f.write("Resumen contextual:\n")
                for palabra, frases in r.get("resumen", {}).items():
                    f.write(f"  - {palabra}:\n")
                    for frase in frases:
                        f.write(f"      * {frase.strip()}\n")
            f.write("\n")

def guardar_excel(resultados, incluir_resumen):
    wb = openpyxl.Workbook()
    ws = wb.active

    encabezados = ["Ruta", "Nombre", "Tipo", "Total coincidencias"] + PALABRAS_CLAVE
    if incluir_resumen:
        encabezados.append("Resumen")
    ws.append(encabezados)

    for r in resultados:
        ruta_completa = os.path.abspath(r["ruta"])
        ruta_directorio = os.path.dirname(ruta_completa)
        nombre_archivo = r["nombre"]

        fila = [ruta_directorio, nombre_archivo, r["tipo"], r["coincidencias"]]
        fila += [r["detalle"].get(p, 0) for p in PALABRAS_CLAVE]

        if incluir_resumen:
            resumen_concat = "\n".join(
                f"{palabra}: " + " | ".join(frases) for palabra, frases in r.get("resumen", {}).items()
            )
            fila.append(resumen_concat)

        ws.append(fila)

    wb.save("informe_documentos.xlsx")

PALABRAS_CLAVE = [
    "requerimientos", "requerimiento", "especificaciones", "funcionalidades", "casos de uso", "caso de uso",
    "arquitectura", "diseño del sistema", "integración", "alcance", "contrato",
    "avance", "informe", "revisión", "desarrollo", "QA", "test", "manual", "características",
    "requisitos", "requisito", "familiarizacion"
]

def cleanup(reprocesar):
    if reprocesar:
        archivos_a_borrar = [
            "analizador.log",
            "informe_documentos.xlsx",
            "informe_documentos.txt",
            "procesados.txt"
        ]
        for archivo in archivos_a_borrar:
            if os.path.exists(archivo):
                try:
                    os.remove(archivo)
                    logger.info(f"🧹 Archivo eliminado: {archivo}")
                except Exception as e:
                    logger.warning(f"❌ No se pudo eliminar {archivo}: {e}")

def main():
    parser = argparse.ArgumentParser(description="Analiza documentos y busca palabras clave.")
    parser.add_argument("--ruta", required=True, help="Ruta de la carpeta que contiene los documentos")
    parser.add_argument("--resumen", action="store_true", help="Incluir resumen contextual")
    parser.add_argument("--reprocesar", action="store_true", help="Limpiar archivos generados previamente")
    args = parser.parse_args()

    if args.reprocesar:
        cleanup(args.reprocesar)

    logger.info(f"⏳ Realizando análisis en: {args.ruta}")

    resultados_previos = cargar_resultados_previos()
    resultados_nuevos = analizar_documentos(args.ruta, args.resumen)
    resultados_totales = resultados_previos + resultados_nuevos

    guardar_txt(resultados_totales, args.resumen)
    guardar_excel(resultados_totales, args.resumen)

    logger.info("✅ Análisis completado.")
    logger.info("🚀 Se actualizaron: informe_documentos.txt, informe_documentos.xlsx")

if __name__ == "__main__":
    main()
