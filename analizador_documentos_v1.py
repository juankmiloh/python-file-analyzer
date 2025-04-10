# SCRIPT V1.0
# Autor: [Tu Nombre]
# Descripción: Este script analiza documentos en varios formatos (PDF, DOCX, XLSX, TXT, CSV)
# y busca palabras clave específicas. Genera un informe en formato TXT y XLSX con los resultados.
# Uso: python script.py --ruta /ruta/a/los/documentos --resumen
# Requisitos: PyPDF2, python-docx, openpyxl
# Licencia: MIT

import os
import re
import PyPDF2
import docx
import openpyxl
import unicodedata
import argparse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor

PALABRAS_CLAVE = [
    "requerimientos", "requerimiento", "especificaciones", "funcionalidades", "casos de uso", "caso de uso",
    "arquitectura", "diseño del sistema", "integración", "alcance", "contrato",
    "avance", "informe", "revisión", "desarrollo", "QA", "test", "manual", "caracteristicas",
    "requisitos", "requisito"
]

def extraer_texto_pdf(path):
    try:
        with open(path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            return " ".join(page.extract_text() or "" for page in reader.pages)
    except Exception as e:
        return f"[Error leyendo PDF: {e}]"

def extraer_texto_docx(path):
    try:
        doc = docx.Document(path)
        return " ".join(paragraph.text for paragraph in doc.paragraphs)
    except Exception as e:
        return f"[Error leyendo DOCX: {e}]"

def extraer_texto_xlsx(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        text = ""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        text += str(cell.value) + " "
        return text
    except Exception as e:
        return f"[Error leyendo XLSX: {e}]"

def normalizar(texto):
    return ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )

def buscar_palabras_clave(texto):
    claves_encontradas = []
    total_coincidencias = 0
    texto_normalizado = normalizar(texto.lower())

    for palabra in PALABRAS_CLAVE:
        palabra_normalizada = normalizar(palabra.lower())
        patron = r'\b' + re.escape(palabra_normalizada) + r'\b'
        matches = re.findall(patron, texto_normalizado)
        if matches:
            claves_encontradas.append(palabra)
            total_coincidencias += len(matches)

    return claves_encontradas, total_coincidencias

def obtener_contexto(texto, palabra, lineas_contexto=2):
    contexto = []
    palabra_normalizada = normalizar(palabra.lower())
    lineas = texto.splitlines()
    for i, linea in enumerate(lineas):
        if re.search(r'\b' + re.escape(palabra_normalizada) + r'\b', normalizar(linea.lower())):
            inicio = max(0, i - lineas_contexto)
            fin = min(len(lineas), i + lineas_contexto + 1)
            fragmento = '\n'.join(lineas[inicio:fin]).strip()
            contexto.append(fragmento)
    return contexto

def procesar_archivo(path, incluir_resumen):
    nombre = os.path.basename(path)
    extension = nombre.lower().split('.')[-1]
    texto = ""
    tipo = extension.upper()

    if extension == "pdf":
        texto = extraer_texto_pdf(path)
    elif extension == "docx":
        texto = extraer_texto_docx(path)
    elif extension == "xlsx":
        texto = extraer_texto_xlsx(path)
    elif extension in ["txt", "csv"]:
        try:
            with open(path, 'r', encoding="utf-8", errors="ignore") as f:
                texto = f.read()
        except Exception as e:
            texto = f"[Error leyendo archivo de texto: {e}]"
    else:
        return None

    if texto.startswith("[Error"):
        return None

    claves_encontradas, total = buscar_palabras_clave(texto)
    if not claves_encontradas:
        return None

    item = {
        "archivo": path,
        "tipo": tipo,
        "palabras_clave": claves_encontradas,
        "coincidencias": total,
        "recomendacion": "Alta" if total > 5 else "Media"
    }

    if incluir_resumen:
        item["resumen"] = {
            palabra: obtener_contexto(texto, palabra)
            for palabra in claves_encontradas
        }

    return item

def analizar_documentos(directorio, incluir_resumen):
    reporte = []
    rutas = []

    for root, _, files in os.walk(directorio):
        for nombre in files:
            rutas.append(os.path.join(root, nombre))

    with ThreadPoolExecutor() as executor:
        resultados = executor.map(lambda path: procesar_archivo(path, incluir_resumen), rutas)
        reporte = [r for r in resultados if r]

    return reporte

def guardar_txt(reporte, incluir_resumen):
    with open("informe_documentos.txt", "w", encoding="utf-8") as f:
        for r in reporte:
            f.write(f"Archivo: {r['archivo']}\n")
            f.write(f"Tipo: {r.get('tipo', 'Desconocido')}\n")
            f.write(f"Palabras clave encontradas: {', '.join(r['palabras_clave'])}\n")
            f.write(f"Número total de coincidencias: {r['coincidencias']}\n")
            f.write(f"Recomendación de revisión: {r['recomendacion']}\n")

            if incluir_resumen and "resumen" in r:
                for palabra, fragmentos in r["resumen"].items():
                    if fragmentos:
                        f.write(f"\n  Contexto para '{palabra}':\n")
                        for i, frag in enumerate(fragmentos[:3]):
                            f.write(f"    {i+1}. ...\n{frag}\n...\n")

            f.write("-" * 40 + "\n\n")

def guardar_excel(reporte, incluir_resumen):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Documentos"

    headers = ["Ruta", "Nombre de archivo", "Tipo", "Palabras Clave", "Coincidencias", "Recomendación"]
    if incluir_resumen:
        headers.append("Resumen Contextual")

    ws.append(headers)

    for r in reporte:
        ruta_completa = r["archivo"]
        ruta_directorio = os.path.dirname(ruta_completa)
        nombre_archivo = os.path.basename(ruta_completa)

        fila = [
            ruta_directorio,
            nombre_archivo,
            r.get("tipo", "Desconocido"),
            ", ".join(r["palabras_clave"]),
            r["coincidencias"],
            r["recomendacion"]
        ]

        if incluir_resumen:
            resumen = ""
            for palabra, fragmentos in r.get("resumen", {}).items():
                if fragmentos:
                    resumen += f"{palabra}:\n"
                    for frag in fragmentos[:2]:
                        resumen += f"  - {frag.strip()[:200]}\n"
            fila.append(resumen.strip())

        ws.append(fila)

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_length + 5, 80)

    wb.save("informe_documentos.xlsx")

def main():
    parser = argparse.ArgumentParser(description="Analiza documentos y busca palabras clave.")
    parser.add_argument("--ruta", required=True, help="Ruta de la carpeta que contiene los documentos")
    parser.add_argument("--resumen", action="store_true", help="Incluir resumen contextual")
    args = parser.parse_args()

    resultado = analizar_documentos(args.ruta, args.resumen)
    guardar_txt(resultado, args.resumen)
    guardar_excel(resultado, args.resumen)

    print("✅ Análisis completado. Se generaron:")
    print(" - informe_documentos.txt")
    print(" - informe_documentos.xlsx")

if __name__ == "__main__":
    main()
